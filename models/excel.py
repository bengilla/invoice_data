import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


# TODO 如果选择了两个公司，应该在Excel显示两个公司
def excel(username: str, company: str, data):
    now = datetime.now()
    date_output = now.strftime("%Y年%m月")
    total_amount = []
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    rows = range(1, 10)
    columns = range(1, 10)
    for row in rows:
        for col in columns:
            ws.cell(row, col).alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

    ws.merge_cells("A1:G1")
    ws.row_dimensions[1].height = 40
    ws["A1"].font = Font(size=14)
    # TODO Excel是否需要名字显示，而不是登入用户名
    ws["A1"] = username + "报销明细表 - " + datetime.now().strftime("%Y年%m月%d日")

    ws.merge_cells("A2:G2")
    ws.row_dimensions[2].height = 20
    ws["A2"].font = Font(size=12)
    ws["A2"] = company

    ws.column_dimensions["B"].width = 17
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["G"].width = 20

    ws["A3"] = "序号"
    ws["B3"] = "时间"
    ws["C3"] = "事项"
    ws["D3"] = "明细/原由"
    ws["E3"] = "金额"
    ws["F3"] = "发票"
    ws["G3"] = "备注"

    index = 1
    column = 4
    for i in data:
        total_amount.append(i.amount)
        ws[f"A{column}"] = index
        ws[f"B{column}"] = i.date
        ws[f"C{column}"] = ""
        ws[f"D{column}"] = i.reason
        ws[f"E{column}"] = i.amount
        ws[f"F{column}"] = "有"
        ws[f"G{column}"] = i.note
        index += 1
        column += 1

    ws.merge_cells(f"A{column}:D{column}")
    # color = ws[f"A{column}:D{column}"]
    color1 = ws[f"A{column}"]
    color1.fill = PatternFill(
        start_color="F0DBAF", end_color="F0DBAF", fill_type="solid"
    )
    color2 = ws[f"E{column}"]
    color2.fill = PatternFill(
        start_color="F0DBAF", end_color="F0DBAF", fill_type="solid"
    )
    ws.row_dimensions[column].height = 40
    ws[f"A{column}"].alignment = Alignment(horizontal="right", vertical="center")
    ws[f"A{column}"] = "合计"
    ws[f"E{column}"] = sum(total_amount)

    # Save the file"
    PATH = os.getcwd() + "/user_file/"
    USER_PATH = PATH + username

    file_name = f"{username}-{date_output}-报销明细表.xlsx"
    file_join = os.path.join(USER_PATH, file_name)
    wb.save(file_join)
    # ----- end excel -----
