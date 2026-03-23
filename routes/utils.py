"""工具函数模块"""

import hashlib
import os
import uuid
from datetime import datetime
from typing import List

import numpy as np
from PIL import Image

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from io import BytesIO


CN_NUMS = ["零", "壹", "贰", "叁", "肆", "伍", "陆", "柒", "捌", "玖"]
CN_INT_RADICE = ["", "拾", "佰", "仟"]
CN_INT_UNITS = ["", "万", "亿", "兆"]
CN_DEC_UNITS = ["角", "分"]


def normalize_unicode(text: str) -> str:
    """规范化Unicode字符 - Kangxi radicals转正常中文"""
    kangxi_map = {"⽂": "文", "⽉": "月", "⽇": "日", "⽕": "火", "⽕": "水"}
    for old, new in kangxi_map.items():
        text = text.replace(old, new)
    return text


def get_file_md5(content: bytes) -> str:
    """计算文件MD5哈希"""
    return hashlib.md5(content).hexdigest()


def to_chinese_upper(amount):
    """将数字金额转换为中文大写"""
    integer_num = int(amount)
    decimal_num = round((amount - integer_num) * 100)
    s = ""
    if integer_num == 0:
        s = CN_NUMS[0]
    else:
        zero_count = 0
        int_str = str(integer_num)
        int_len = len(int_str)
        for i in range(int_len):
            n = int(int_str[i])
            p = int_len - i - 1
            quotient = p // 4
            modulus = p % 4
            if n == 0:
                zero_count += 1
            else:
                if zero_count > 0:
                    s += CN_NUMS[0]
                zero_count = 0
                s += CN_NUMS[n] + CN_INT_RADICE[modulus]
            if modulus == 0 and zero_count < 4:
                s += CN_INT_UNITS[quotient]
    s += "元"
    if decimal_num == 0:
        s += "整"
    else:
        jiao = decimal_num // 10
        fen = decimal_num % 10
        if jiao > 0:
            s += CN_NUMS[jiao] + CN_DEC_UNITS[0]
        if fen > 0:
            s += CN_NUMS[fen] + CN_DEC_UNITS[1]
    return "人民币 " + s


def generate_excel(meta, invoices, total_amount, total_invoice, total_other_invoice):
    """生成Excel报销单"""
    wb = Workbook()
    ws = wb.active
    ws.title = "费用报销单"

    ft_title = Font(name="微软雅黑", size=18, bold=True)
    ft_label = Font(name="微软雅黑", size=11)
    ft_data = Font(name="微软雅黑", size=11)
    ft_header = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    ft_total = Font(name="微软雅黑", size=12, bold=True)

    al_center = Alignment(horizontal="center", vertical="center")
    al_left = Alignment(horizontal="left", vertical="center")
    al_right = Alignment(horizontal="right", vertical="center")

    s_thin = Side(style="thin", color="000000")
    s_none = Side(style=None)
    border_table = Border(left=s_thin, right=s_thin, top=s_thin, bottom=s_thin)

    fill_header = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    fill_light = PatternFill(
        start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"
    )

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 19
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 20

    ws.row_dimensions[1].height = 45
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 28
    ws.row_dimensions[4].height = 28
    ws.row_dimensions[5].height = 10
    ws.row_dimensions[6].height = 30

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "费 用 报 销 单"
    c.font = ft_title
    c.alignment = al_center

    ws["A2"].value = "报销部门"
    ws["A2"].font = ft_label
    ws["A2"].alignment = al_right
    ws.merge_cells("B2:B2")
    ws["B2"].value = meta.get("department", "")
    ws["B2"].font = ft_data
    ws["B2"].alignment = al_left
    ws["B2"].border = Border(bottom=Side(style="thin", color="000000"))
    ws["D2"].value = "报销日期"
    ws["D2"].font = ft_label
    ws["D2"].alignment = al_right
    ws["E2"].value = meta.get("date", "")
    ws["E2"].font = ft_data
    ws["E2"].alignment = al_left
    ws["E2"].border = Border(bottom=Side(style="thin", color="000000"))

    ws["A3"].value = "报销人"
    ws["A3"].font = ft_label
    ws["A3"].alignment = al_right
    ws["B3"].value = meta.get("name", "")
    ws["B3"].font = ft_data
    ws["B3"].alignment = al_left
    ws["B3"].border = Border(bottom=Side(style="thin", color="000000"))
    ws["D3"].value = "所属岗位"
    ws["D3"].font = ft_label
    ws["D3"].alignment = al_right
    ws["E3"].value = meta.get("position", "")
    ws["E3"].font = ft_data
    ws["E3"].alignment = al_left
    ws["E3"].border = Border(bottom=Side(style="thin", color="000000"))

    ws["A4"].value = "事由"
    ws["A4"].font = ft_label
    ws["A4"].alignment = al_right
    ws.merge_cells("B4:E4")
    ws["B4"].value = meta.get("use", "")
    ws["B4"].font = ft_data
    ws["B4"].alignment = al_left
    ws["B4"].border = Border(bottom=Side(style="thin", color="000000"))
    for col in ["C", "D", "E"]:
        ws[f"{col}4"].border = Border(bottom=Side(style="thin", color="000000"))

    headers = [
        ("A6", "序号"),
        ("B6", "日期"),
        ("C6", "事项"),
        ("D6", "明细/原由"),
        ("E6", "金额（元）"),
        ("F6", "备注"),
    ]
    for cell_ref, val in headers:
        c = ws[cell_ref]
        c.value = val
        c.font = ft_header
        c.alignment = al_center
        c.fill = fill_header
        c.border = border_table

    invoice_count = len(invoices)
    for idx in range(invoice_count):
        row = 7 + idx
        ws.row_dimensions[row].height = 28
        inv = invoices[idx]

        c = ws[f"A{row}"]
        c.value = idx + 1
        c.font = ft_data
        c.alignment = al_center
        c.border = border_table

        date_val = inv.get("date", "") if inv else ""
        if date_val and "-" in date_val:
            parts = date_val.split("-")
            if len(parts) == 3:
                date_val = f"{parts[0]}年{int(parts[1])}月{int(parts[2])}日"
        c = ws[f"B{row}"]
        c.value = date_val
        c.font = ft_data
        c.alignment = al_center
        c.border = border_table

        c = ws[f"C{row}"]
        c.value = inv.get("company", "") if inv else ""
        c.font = ft_data
        c.alignment = al_left
        c.border = border_table

        c = ws[f"D{row}"]
        c.value = inv.get("reason", "") if inv else ""
        c.font = ft_data
        c.alignment = al_left
        c.border = border_table

        c = ws[f"E{row}"]
        if inv:
            amt = inv.get("amount", 0)
            c.value = f"¥{amt:,.2f}"
        else:
            c.value = ""
        c.font = ft_data
        c.alignment = al_right
        c.border = border_table

        c = ws[f"F{row}"]
        c.value = inv.get("remark", "") if inv else ""
        c.font = ft_data
        c.alignment = al_left
        c.border = border_table

        if idx % 2 == 0 and inv:
            for col in ["A", "B", "C", "D", "E", "F"]:
                ws[f"{col}{row}"].fill = fill_light

    total_row = 7 + invoice_count
    ws.row_dimensions[total_row].height = 30
    ws.merge_cells(f"A{total_row}:C{total_row}")
    c = ws[f"A{total_row}"]
    c.value = "合  计"
    c.font = ft_header
    c.alignment = al_center
    c.fill = fill_header
    c.border = border_table
    ws[f"B{total_row}"].border = border_table
    ws[f"C{total_row}"].border = border_table

    ws.merge_cells(f"D{total_row}:F{total_row}")
    c = ws[f"D{total_row}"]
    c.value = f"¥{total_amount:,.2f}"
    c.font = Font(name="微软雅黑", size=13, bold=True, color="C00000")
    c.alignment = al_right
    c.border = border_table
    ws[f"E{total_row}"].border = border_table
    ws[f"F{total_row}"].border = border_table

    amount_row = total_row + 1
    ws.row_dimensions[amount_row].height = 28
    ws.merge_cells(f"A{amount_row}:F{amount_row}")
    c = ws[f"A{amount_row}"]
    c.value = f"金额大写（人民币）：{to_chinese_upper(total_amount)}"
    c.font = Font(name="微软雅黑", size=11, bold=True)
    c.alignment = al_left

    empty_row = amount_row + 1
    ws.row_dimensions[empty_row].height = 10

    attach_row = empty_row + 1
    ws.row_dimensions[attach_row].height = 26
    ws.merge_cells(f"A{attach_row}:F{attach_row}")
    ws[f"A{attach_row}"].value = "附  件  说  明"
    ws[f"A{attach_row}"].font = Font(name="微软雅黑", size=11, bold=True)
    ws[f"A{attach_row}"].alignment = al_center

    invoice_count_row = attach_row + 1
    ws.row_dimensions[invoice_count_row].height = 26
    ws[f"A{invoice_count_row}"].value = "发票共"
    ws[f"A{invoice_count_row}"].font = ft_label
    ws[f"A{invoice_count_row}"].alignment = al_right
    ws[f"B{invoice_count_row}"].value = total_invoice
    ws[f"B{invoice_count_row}"].font = ft_data
    ws[f"B{invoice_count_row}"].alignment = al_center
    ws[f"B{invoice_count_row}"].border = Border(bottom=Side(style="thin", color="000000"))
    ws.merge_cells(f"C{invoice_count_row}:F{invoice_count_row}")
    ws[f"C{invoice_count_row}"].value = "张"
    ws[f"C{invoice_count_row}"].font = ft_label
    ws[f"C{invoice_count_row}"].alignment = al_left

    other_row = invoice_count_row + 1
    ws.row_dimensions[other_row].height = 26
    ws[f"A{other_row}"].value = "其他单据共"
    ws[f"A{other_row}"].font = ft_label
    ws[f"A{other_row}"].alignment = al_right
    ws[f"B{other_row}"].value = total_other_invoice
    ws[f"B{other_row}"].font = ft_data
    ws[f"B{other_row}"].alignment = al_center
    ws[f"B{other_row}"].border = Border(bottom=Side(style="thin", color="000000"))
    ws.merge_cells(f"C{other_row}:F{other_row}")
    ws[f"C{other_row}"].value = "张"
    ws[f"C{other_row}"].font = ft_label
    ws[f"C{other_row}"].alignment = al_left

    empty_row2 = other_row + 1
    ws.row_dimensions[empty_row2].height = 10

    approval_row = empty_row2 + 1
    ws.row_dimensions[approval_row].height = 30
    ws.merge_cells(f"A{approval_row}:F{approval_row}")
    ws[f"A{approval_row}"].value = "审  批  签  字"
    ws[f"A{approval_row}"].font = Font(name="微软雅黑", size=11, bold=True)
    ws[f"A{approval_row}"].alignment = al_center

    sign_items = ["报销人签字", "部门负责人", "财务审核", "公司负责人"]
    for i, text in enumerate(sign_items):
        row = approval_row + 1 + i
        ws.row_dimensions[row].height = 36
        ws[f"A{row}"].value = text + "："
        ws[f"A{row}"].font = ft_label
        ws[f"A{row}"].alignment = al_right
        ws.merge_cells(f"B{row}:F{row}")
        ws[f"B{row}"].border = Border(bottom=Side(style="thin", color="000000"))
        ws[f"C{row}"].border = Border(bottom=Side(style="thin", color="000000"))
        ws[f"D{row}"].border = Border(bottom=Side(style="thin", color="000000"))
        ws[f"E{row}"].border = Border(bottom=Side(style="thin", color="000000"))
        ws[f"F{row}"].border = Border(bottom=Side(style="thin", color="000000"))

    last_row = approval_row + 4
    ws.print_area = f"A1:F{last_row}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


OCR_ENGINE = None


def get_ocr_engine():
    """获取OCR引擎（单例模式）"""
    global OCR_ENGINE
    if OCR_ENGINE is None:
        from rapidocr_onnxruntime import RapidOCR

        OCR_ENGINE = RapidOCR()
    return OCR_ENGINE


def parse_invoice_image(file_path: str) -> dict:
    """使用OCR解析图片发票"""
    import re

    result = {
        "file_name": os.path.basename(file_path),
        "date": None,
        "year": datetime.now().year,
        "month": datetime.now().month,
        "company": "",
        "buyer": "未知",
        "amount": 0.0,
        "invoice_no": "",
        "success": False,
    }

    try:
        ocr = get_ocr_engine()
        img = Image.open(file_path)
        img_array = np.array(img)

        ocr_result, elapse = ocr(img_array)

        if not ocr_result:
            result["error"] = "未识别到文字"
            return result

        text_lines = [line[1] for line in ocr_result]
        text = "\n".join(text_lines)

        date_match = re.search(r"(\d{4})年(\d{1,2})[⽉月](\d{1,2})[⽇日]", text)
        if date_match:
            result["year"] = int(date_match.group(1))
            result["month"] = int(date_match.group(2))
            result["date"] = (
                f"{date_match.group(1)}-{date_match.group(2).zfill(2)}-{date_match.group(3).zfill(2)}"
            )

        double_match = re.findall(r"名\s*称[：:]\s*([^\s\n]{2,50})", text)
        if len(double_match) >= 1:
            result["buyer"] = double_match[0].strip()
        else:
            lines = text.split("\n")
            for line in lines:
                parts = line.split()
                company_candidates = [p for p in parts if len(p) >= 4 and "公司" in p]
                if len(company_candidates) >= 1:
                    result["buyer"] = company_candidates[0]
                    break

        if result["buyer"].startswith("名称：") or result["buyer"].startswith("名称:"):
            result["buyer"] = result["buyer"][3:]

        result["buyer"] = normalize_unicode(result["buyer"])

        amount_match = re.search(r"[¥￥]([0-9,]+\.?\d*)", text)
        if amount_match:
            amount_str = amount_match.group(1).replace(",", "")
            result["amount"] = float(amount_str)

        no_match = re.search(r"发票号码[：:]*\s*(\d+)", text)
        if no_match:
            result["invoice_no"] = no_match.group(1)
        else:
            no20_match = re.search(r"\b(\d{20})\b", text)
            if no20_match:
                result["invoice_no"] = no20_match.group(1)

        result["success"] = True

    except Exception as e:
        result["error"] = str(e)

    return result


def parse_invoice_pdf(file_path: str) -> dict:
    """解析单个发票PDF"""
    import re

    result = {
        "file_name": os.path.basename(file_path),
        "date": None,
        "year": datetime.now().year,
        "month": datetime.now().month,
        "company": "",
        "buyer": "未知",
        "amount": 0.0,
        "invoice_no": "",
        "success": False,
    }

    try:
        with pdfplumber.open(file_path) as pdf:
            text = ""
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"

        date_match = re.search(r"(\d{4})年(\d{1,2})[⽉月](\d{1,2})[⽇日]", text)
        if date_match:
            result["year"] = int(date_match.group(1))
            result["month"] = int(date_match.group(2))
            result["date"] = (
                f"{date_match.group(1)}-{date_match.group(2).zfill(2)}-{date_match.group(3).zfill(2)}"
            )

        double_match = re.findall(r"名\s*称[：:]\s*([^\s\n]{2,50})", text)
        if len(double_match) >= 1:
            result["buyer"] = double_match[0].strip()
        else:
            lines = text.split("\n")
            for line in lines:
                parts = line.split()
                company_candidates = [p for p in parts if len(p) >= 4 and "公司" in p]
                if len(company_candidates) >= 1:
                    result["buyer"] = company_candidates[0]
                    break

        if result["buyer"].startswith("名称：") or result["buyer"].startswith("名称:"):
            result["buyer"] = result["buyer"][3:]

        result["buyer"] = normalize_unicode(result["buyer"])

        amount_match = re.search(r"[¥￥]([0-9,]+\.?\d*)", text)
        if amount_match:
            result["amount"] = float(amount_match.group(1))

        no_match = re.search(r"发票号码[：:]*\s*(\d+)", text)
        if no_match:
            result["invoice_no"] = no_match.group(1)
        else:
            no20_match = re.search(r"\b(\d{20})\b", text)
            if no20_match:
                result["invoice_no"] = no20_match.group(1)

        result["success"] = True

    except Exception as e:
        result["error"] = str(e)

    return result
