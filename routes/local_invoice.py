"""本地发票处理（无需数据库）"""

import os
import re
import base64
import json
import zipfile
import io
import uuid
import numpy as np
from datetime import datetime
from typing import List

import pdfplumber
from PIL import Image
from fastapi import APIRouter, UploadFile, File
from fastapi.responses import JSONResponse

local_invoice_routes = APIRouter()

TEMP_SESSIONS = {}

CN_NUMS = ["零", "壹", "贰", "叁", "肆", "伍", "陆", "柒", "捌", "玖"]
CN_INT_RADICE = ["", "拾", "佰", "仟"]
CN_INT_UNITS = ["", "万", "亿", "兆"]
CN_DEC_UNITS = ["角", "分"]


def to_chinese_upper(amount):
    integer_num = int(amount)
    decimal_num = round((amount - integer_num) * 100)
    s = ""
    if integer_num == 0:
        s = CN_NUMS[0]
    else:
        zero_count = 0
        int_str = str(integer_num)
        int_len = len(int_str)
        for i in range(int_len):
            n = int(int_str[i])
            p = int_len - i - 1
            quotient = p // 4
            modulus = p % 4
            if n == 0:
                zero_count += 1
            else:
                if zero_count > 0:
                    s += CN_NUMS[0]
                zero_count = 0
                s += CN_NUMS[n] + CN_INT_RADICE[modulus]
            if modulus == 0 and zero_count < 4:
                s += CN_INT_UNITS[quotient]
    s += "元"
    if decimal_num == 0:
        s += "整"
    else:
        jiao = decimal_num // 10
        fen = decimal_num % 10
        if jiao > 0:
            s += CN_NUMS[jiao] + CN_DEC_UNITS[0]
        if fen > 0:
            s += CN_NUMS[fen] + CN_DEC_UNITS[1]
    return "人民币 " + s


def generate_excel(meta, invoices, total_amount, total_invoice, total_other_invoice):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "费用报销单"

    # 字体
    ft_title = Font(name="微软雅黑", size=18, bold=True)
    ft_label = Font(name="微软雅黑", size=11)
    ft_data = Font(name="微软雅黑", size=11)
    ft_header = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    ft_total = Font(name="微软雅黑", size=12, bold=True)

    # 对齐
    al_center = Alignment(horizontal="center", vertical="center")
    al_left = Alignment(horizontal="left", vertical="center")
    al_right = Alignment(horizontal="right", vertical="center")

    # 边框
    s_thin = Side(style="thin", color="000000")
    s_none = Side(style=None)
    border_table = Border(left=s_thin, right=s_thin, top=s_thin, bottom=s_thin)

    # 填充
    fill_header = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    fill_light = PatternFill(
        start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"
    )

    # 列宽
    ws.column_dimensions["A"].width = 15  # 序号
    ws.column_dimensions["B"].width = 19  # 日期
    ws.column_dimensions["C"].width = 16  # 事项
    ws.column_dimensions["D"].width = 40  # 明细/原由
    ws.column_dimensions["E"].width = 14  # 金额
    ws.column_dimensions["F"].width = 20  # 备注

    # 行高
    ws.row_dimensions[1].height = 45
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 28
    ws.row_dimensions[4].height = 28
    ws.row_dimensions[5].height = 10
    ws.row_dimensions[6].height = 30

    # ============ Row 1: 标题 ============
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "费 用 报 销 单"
    c.font = ft_title
    c.alignment = al_center

    # ============ Row 2: 报销部门 + 日期 ============
    ws["A2"].value = "报销部门"
    ws["A2"].font = ft_label
    ws["A2"].alignment = al_right
    ws.merge_cells("B2:B2")
    ws["B2"].value = meta.get("department", "")
    ws["B2"].font = ft_data
    ws["B2"].alignment = al_left
    ws["B2"].border = Border(bottom=Side(style="thin", color="000000"))
    ws["D2"].value = "报销日期"
    ws["D2"].font = ft_label
    ws["D2"].alignment = al_right
    ws["E2"].value = meta.get("date", "")
    ws["E2"].font = ft_data
    ws["E2"].alignment = al_left
    ws["E2"].border = Border(bottom=Side(style="thin", color="000000"))

    # ============ Row 3: 报销人 + 岗位 ============
    ws["A3"].value = "报销人"
    ws["A3"].font = ft_label
    ws["A3"].alignment = al_right
    ws["B3"].value = meta.get("name", "")
    ws["B3"].font = ft_data
    ws["B3"].alignment = al_left
    ws["B3"].border = Border(bottom=Side(style="thin", color="000000"))
    ws["D3"].value = "所属岗位"
    ws["D3"].font = ft_label
    ws["D3"].alignment = al_right
    ws["E3"].value = meta.get("position", "")
    ws["E3"].font = ft_data
    ws["E3"].alignment = al_left
    ws["E3"].border = Border(bottom=Side(style="thin", color="000000"))

    # ============ Row 4: 事由 ============
    ws["A4"].value = "事由"
    ws["A4"].font = ft_label
    ws["A4"].alignment = al_right
    ws.merge_cells("B4:E4")
    ws["B4"].value = meta.get("use", "")
    ws["B4"].font = ft_data
    ws["B4"].alignment = al_left
    ws["B4"].border = Border(bottom=Side(style="thin", color="000000"))
    for col in ["C", "D", "E"]:
        ws[f"{col}4"].border = Border(bottom=Side(style="thin", color="000000"))

    # ============ Row 5: 空行 ============

    # ============ Row 6: 表头 ============
    headers = [
        ("A6", "序号"),
        ("B6", "日期"),
        ("C6", "事项"),
        ("D6", "明细/原由"),
        ("E6", "金额（元）"),
        ("F6", "备注"),
    ]
    for cell_ref, val in headers:
        c = ws[cell_ref]
        c.value = val
        c.font = ft_header
        c.alignment = al_center
        c.fill = fill_header
        c.border = border_table

    # ============ 数据行 ============
    invoice_count = len(invoices)
    for idx in range(invoice_count):
        row = 7 + idx
        ws.row_dimensions[row].height = 28
        inv = invoices[idx]

        # 序号
        c = ws[f"A{row}"]
        c.value = idx + 1
        c.font = ft_data
        c.alignment = al_center
        c.border = border_table

        # 日期
        date_val = inv.get("date", "") if inv else ""
        if date_val and "-" in date_val:
            parts = date_val.split("-")
            if len(parts) == 3:
                date_val = f"{parts[0]}年{int(parts[1])}月{int(parts[2])}日"
        c = ws[f"B{row}"]
        c.value = date_val
        c.font = ft_data
        c.alignment = al_center
        c.border = border_table

        # 事项
        c = ws[f"C{row}"]
        c.value = inv.get("company", "") if inv else ""
        c.font = ft_data
        c.alignment = al_left
        c.border = border_table

        # 明细/原由
        c = ws[f"D{row}"]
        c.value = inv.get("reason", "") if inv else ""
        c.font = ft_data
        c.alignment = al_left
        c.border = border_table

        # 金额 (带¥)
        c = ws[f"E{row}"]
        if inv:
            amt = inv.get("amount", 0)
            c.value = f"¥{amt:,.2f}"
        else:
            c.value = ""
        c.font = ft_data
        c.alignment = al_right
        c.border = border_table

        # 备注
        c = ws[f"F{row}"]
        c.value = inv.get("remark", "") if inv else ""
        c.font = ft_data
        c.alignment = al_left
        c.border = border_table

        # 隔行变色
        if idx % 2 == 0 and inv:
            for col in ["A", "B", "C", "D", "E", "F"]:
                ws[f"{col}{row}"].fill = fill_light

    # ============ 合计 ============
    total_row = 7 + invoice_count
    ws.row_dimensions[total_row].height = 30
    ws.merge_cells(f"A{total_row}:C{total_row}")
    c = ws[f"A{total_row}"]
    c.value = "合  计"
    c.font = ft_header
    c.alignment = al_center
    c.fill = fill_header
    c.border = border_table
    ws[f"B{total_row}"].border = border_table
    ws[f"C{total_row}"].border = border_table

    ws.merge_cells(f"D{total_row}:F{total_row}")
    c = ws[f"D{total_row}"]
    c.value = f"¥{total_amount:,.2f}"
    c.font = Font(name="微软雅黑", size=13, bold=True, color="C00000")
    c.alignment = al_right
    c.border = border_table
    ws[f"E{total_row}"].border = border_table
    ws[f"F{total_row}"].border = border_table

    # ============ 金额大写 ============
    amount_row = total_row + 1
    ws.row_dimensions[amount_row].height = 28
    ws.merge_cells(f"A{amount_row}:F{amount_row}")
    c = ws[f"A{amount_row}"]
    c.value = f"金额大写（人民币）：{to_chinese_upper(total_amount)}"
    c.font = Font(name="微软雅黑", size=11, bold=True)
    c.alignment = al_left

    # ============ 空行 ============
    empty_row = amount_row + 1
    ws.row_dimensions[empty_row].height = 10

    # ============ 附件说明 ============
    attach_row = empty_row + 1
    ws.row_dimensions[attach_row].height = 26
    ws.merge_cells(f"A{attach_row}:F{attach_row}")
    ws[f"A{attach_row}"].value = "附  件  说  明"
    ws[f"A{attach_row}"].font = Font(name="微软雅黑", size=11, bold=True)
    ws[f"A{attach_row}"].alignment = al_center

    # ============ 发票共__张 ============
    invoice_count_row = attach_row + 1
    ws.row_dimensions[invoice_count_row].height = 26
    ws[f"A{invoice_count_row}"].value = "发票共"
    ws[f"A{invoice_count_row}"].font = ft_label
    ws[f"A{invoice_count_row}"].alignment = al_right
    ws[f"B{invoice_count_row}"].value = total_invoice
    ws[f"B{invoice_count_row}"].font = ft_data
    ws[f"B{invoice_count_row}"].alignment = al_center
    ws[f"B{invoice_count_row}"].border = Border(bottom=Side(style="thin", color="000000"))
    ws.merge_cells(f"C{invoice_count_row}:F{invoice_count_row}")
    ws[f"C{invoice_count_row}"].value = "张"
    ws[f"C{invoice_count_row}"].font = ft_label
    ws[f"C{invoice_count_row}"].alignment = al_left

    # ============ 其他单据共__张 ============
    other_row = invoice_count_row + 1
    ws.row_dimensions[other_row].height = 26
    ws[f"A{other_row}"].value = "其他单据共"
    ws[f"A{other_row}"].font = ft_label
    ws[f"A{other_row}"].alignment = al_right
    ws[f"B{other_row}"].value = total_other_invoice
    ws[f"B{other_row}"].font = ft_data
    ws[f"B{other_row}"].alignment = al_center
    ws[f"B{other_row}"].border = Border(bottom=Side(style="thin", color="000000"))
    ws.merge_cells(f"C{other_row}:F{other_row}")
    ws[f"C{other_row}"].value = "张"
    ws[f"C{other_row}"].font = ft_label
    ws[f"C{other_row}"].alignment = al_left

    # ============ 空行 ============
    empty_row2 = other_row + 1
    ws.row_dimensions[empty_row2].height = 10

    # ============ 审批签字 ============
    approval_row = empty_row2 + 1
    ws.row_dimensions[approval_row].height = 30
    ws.merge_cells(f"A{approval_row}:F{approval_row}")
    ws[f"A{approval_row}"].value = "审  批  签  字"
    ws[f"A{approval_row}"].font = Font(name="微软雅黑", size=11, bold=True)
    ws[f"A{approval_row}"].alignment = al_center

    # ============ 签字栏（竖排，带下划线） ============
    sign_items = ["报销人签字", "部门负责人", "财务审核", "公司负责人"]
    for i, text in enumerate(sign_items):
        row = approval_row + 1 + i
        ws.row_dimensions[row].height = 36
        # 标签
        ws[f"A{row}"].value = text + "："
        ws[f"A{row}"].font = ft_label
        ws[f"A{row}"].alignment = al_right
        # 签名区域（下划线）
        ws.merge_cells(f"B{row}:F{row}")
        ws[f"B{row}"].border = Border(bottom=Side(style="thin", color="000000"))
        ws[f"C{row}"].border = Border(bottom=Side(style="thin", color="000000"))
        ws[f"D{row}"].border = Border(bottom=Side(style="thin", color="000000"))
        ws[f"E{row}"].border = Border(bottom=Side(style="thin", color="000000"))
        ws[f"F{row}"].border = Border(bottom=Side(style="thin", color="000000"))

    # 打印设置
    last_row = approval_row + 4
    ws.print_area = f"A1:F{last_row}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


OCR_ENGINE = None


def get_ocr_engine():
    global OCR_ENGINE
    if OCR_ENGINE is None:
        from rapidocr_onnxruntime import RapidOCR

        OCR_ENGINE = RapidOCR()
    return OCR_ENGINE


def parse_invoice_image(file_path: str) -> dict:
    """使用OCR解析图片发票"""
    result = {
        "file_name": os.path.basename(file_path),
        "date": None,
        "year": datetime.now().year,
        "month": datetime.now().month,
        "company": "",
        "buyer": "未知",
        "amount": 0.0,
        "invoice_no": "",
        "success": False,
    }

    try:
        ocr = get_ocr_engine()
        img = Image.open(file_path)
        img_array = np.array(img)

        ocr_result, elapse = ocr(img_array)

        if not ocr_result:
            result["error"] = "未识别到文字"
            return result

        text_lines = [line[1] for line in ocr_result]
        text = "\n".join(text_lines)

        date_match = re.search(r"(\d{4})年(\d{1,2})[⽉月](\d{1,2})[⽇日]", text)
        if date_match:
            result["year"] = int(date_match.group(1))
            result["month"] = int(date_match.group(2))
            result["date"] = (
                f"{date_match.group(1)}-{date_match.group(2).zfill(2)}-{date_match.group(3).zfill(2)}"
            )

        double_match = re.findall(r"名\s*称[：:]\s*([^\s\n]{2,50})", text)
        if len(double_match) >= 1:
            result["buyer"] = double_match[0].strip()
        else:
            lines = text.split("\n")
            for line in lines:
                parts = line.split()
                company_candidates = [p for p in parts if len(p) >= 4 and "公司" in p]
                if len(company_candidates) >= 1:
                    result["buyer"] = company_candidates[0]
                    break

        # 移除"名称："前缀
        if result["buyer"].startswith("名称：") or result["buyer"].startswith("名称:"):
            result["buyer"] = result["buyer"][3:]

        # 规范化Unicode字符 - Kangxi radicals转正常中文
        kangxi_map = {"⽂": "文", "⽉": "月", "⽇": "日", "⽕": "火", "⽕": "水"}
        for old, new in kangxi_map.items():
            result["buyer"] = result["buyer"].replace(old, new)

        # 提取金额 - 支持普通发票和火车票(票价:￥xxx)
        amount_match = re.search(r"[¥￥]([0-9,]+\.?\d*)", text)
        if amount_match:
            amount_str = amount_match.group(1).replace(",", "")
            result["amount"] = float(amount_str)

        no_match = re.search(r"发票号码[：:]*\s*(\d+)", text)
        if no_match:
            result["invoice_no"] = no_match.group(1)
        else:
            no20_match = re.search(r"\b(\d{20})\b", text)
            if no20_match:
                result["invoice_no"] = no20_match.group(1)

        result["success"] = True

    except Exception as e:
        result["error"] = str(e)

    return result


def parse_invoice_pdf(file_path: str) -> dict:
    """解析单个发票PDF"""
    result = {
        "file_name": os.path.basename(file_path),
        "date": None,
        "year": datetime.now().year,
        "month": datetime.now().month,
        "company": "",
        "buyer": "未知",
        "amount": 0.0,
        "invoice_no": "",
        "success": False,
    }

    try:
        with pdfplumber.open(file_path) as pdf:
            text = ""
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"

        # 提取日期 (2026年03月10日) - 支持月(月/⽉)和日(日/⽇)的多种Unicode字符
        date_match = re.search(r"(\d{4})年(\d{1,2})[⽉月](\d{1,2})[⽇日]", text)
        if date_match:
            result["year"] = int(date_match.group(1))
            result["month"] = int(date_match.group(2))
            result["date"] = (
                f"{date_match.group(1)}-{date_match.group(2).zfill(2)}-{date_match.group(3).zfill(2)}"
            )

        # 提取公司名称 - 名称：xxx 名称：xxx 格式，取第一个为购方
        double_match = re.findall(r"名\s*称[：:]\s*([^\s\n]{2,50})", text)
        if len(double_match) >= 1:
            result["buyer"] = double_match[0].strip()
        else:
            lines = text.split("\n")
            for line in lines:
                parts = line.split()
                company_candidates = [p for p in parts if len(p) >= 4 and "公司" in p]
                if len(company_candidates) >= 1:
                    result["buyer"] = company_candidates[0]
                    break

        # 移除"名称："前缀
        if result["buyer"].startswith("名称：") or result["buyer"].startswith("名称:"):
            result["buyer"] = result["buyer"][3:]

        # 规范化Unicode字符 - Kangxi radicals转正常中文
        kangxi_map = {"⽂": "文", "⽉": "月", "⽇": "日", "⽕": "火", "⽕": "水"}
        for old, new in kangxi_map.items():
            result["buyer"] = result["buyer"].replace(old, new)

        # 提取金额 - 支持普通发票和火车票(票价:￥xxx)
        amount_match = re.search(r"[¥￥]([0-9,]+\.?\d*)", text)
        if amount_match:
            result["amount"] = float(amount_match.group(1))

        # 提取发票号码
        no_match = re.search(r"发票号码[：:]*\s*(\d+)", text)
        if no_match:
            result["invoice_no"] = no_match.group(1)
        else:
            no20_match = re.search(r"\b(\d{20})\b", text)
            if no20_match:
                result["invoice_no"] = no20_match.group(1)

        result["success"] = True

    except Exception as e:
        result["error"] = str(e)

    return result


@local_invoice_routes.post("/api/parse-invoices")
async def parse_invoices(files: List[UploadFile] = File(...)):
    """解析上传的发票文件（PDF或图片）"""
    import hashlib

    all_invoices = []
    months_data = {}
    duplicate_files = []
    processed_filenames = set()
    processed_hashes = set()

    session_id = str(uuid.uuid4())[:8]
    temp_dir = f"/tmp/invoice_{session_id}"
    os.makedirs(temp_dir, exist_ok=True)
    TEMP_SESSIONS[session_id] = temp_dir

    print(f"Received {len(files)} files, session: {session_id}")

    supported_extensions = {".pdf", ".png", ".jpg", ".jpeg", ".bmp", ".gif"}

    try:
        for file in files:
            print(f"Processing file: {file.filename}")

            ext = os.path.splitext(file.filename)[1].lower()
            if ext not in supported_extensions:
                print(f"Skipping unsupported: {file.filename}")
                continue

            # 读取文件内容
            content = await file.read()

            # 计算文件哈希
            file_hash = hashlib.md5(content).hexdigest()

            # 检查重复 - 文件名重复或内容重复
            if file.filename in processed_filenames:
                duplicate_files.append(file.filename)
                print(f"Duplicate filename skipped: {file.filename}")
                continue

            if file_hash in processed_hashes:
                duplicate_files.append(file.filename)
                print(f"Duplicate content skipped: {file.filename}")
                continue

            processed_filenames.add(file.filename)
            processed_hashes.add(file_hash)

            file_path = os.path.join(temp_dir, file.filename)

            with open(file_path, "wb") as buffer:
                buffer.write(content)
            print(f"Saved: {file_path}")

            if ext == ".pdf":
                invoice = parse_invoice_pdf(file_path)
            else:
                invoice = parse_invoice_image(file_path)

            if invoice["success"]:
                all_invoices.append(invoice)

                key = f"{invoice['year']}-{invoice['month']:02d}"
                if key not in months_data:
                    months_data[key] = {
                        "year": invoice["year"],
                        "month": invoice["month"],
                        "invoices": [],
                        "total": 0.0,
                    }
                months_data[key]["invoices"].append(invoice)
                months_data[key]["total"] += invoice["amount"]

        return JSONResponse(
            {
                "success": True,
                "session_id": session_id,
                "invoices": all_invoices,
                "months": list(months_data.values()),
                "total_amount": sum(inv["amount"] for inv in all_invoices),
                "duplicates": duplicate_files,
            }
        )

    except Exception as e:
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)


@local_invoice_routes.post("/api/download-month-zip")
async def download_month_zip(data: dict):
    """打包月度发票原件ZIP"""
    from fastapi.responses import StreamingResponse
    import zipfile
    from io import BytesIO

    session_id = data.get("session_id")
    year = data.get("year")
    month = data.get("month")
    invoices = data.get("invoices", [])
    total = data.get("total", 0)
    meta = data.get("meta", {})

    if not session_id or session_id not in TEMP_SESSIONS:
        return JSONResponse(
            {"success": False, "error": "Session expired"}, status_code=400
        )

    temp_dir = TEMP_SESSIONS[session_id]

    buffer = BytesIO()

    try:
        with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            for inv in invoices:
                file_name = inv.get("file_name")
                if not file_name:
                    continue

                file_path = os.path.join(temp_dir, file_name)

                if os.path.exists(file_path):
                    zip_file.write(file_path, file_name)

            # 生成Excel报销单
            try:
                meta["total"] = total
                meta["chinese_price"] = f"金额大写：{to_chinese_upper(total)}"
                meta["total_invoice"] = len(invoices)
                meta["total_other_invoice"] = data.get("total_other_invoice", 0)
                excel_buf = generate_excel(
                    meta, invoices, total, len(invoices), meta["total_other_invoice"]
                )
                zip_file.writestr("费用报销单.xlsx", excel_buf.read())
                print("Excel added to ZIP successfully")
            except Exception as e:
                print(f"Excel generation failed: {e}")
                import traceback

                traceback.print_exc()

        buffer.seek(0)

        filename = f"{year}-{total:.2f}.zip"
        return StreamingResponse(
            buffer,
            media_type="application/zip",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
        )
    finally:
        pass


@local_invoice_routes.post("/api/download-all-zip")
async def download_all_zip(data: dict):
    """打包全部发票ZIP，按月份分文件夹"""
    from fastapi.responses import StreamingResponse
    import zipfile
    from io import BytesIO

    session_id = data.get("session_id")
    months = data.get("months", [])
    total_amount = data.get("total_amount", 0)
    meta = data.get("meta", {})

    if not session_id or session_id not in TEMP_SESSIONS:
        return JSONResponse(
            {"success": False, "error": "Session expired"}, status_code=400
        )

    temp_dir = TEMP_SESSIONS[session_id]
    first_year = months[0].get("year") if months else "2026"

    # 收集全部发票
    all_invoices = []
    for m in months:
        all_invoices.extend(m.get("invoices", []))

    buffer = BytesIO()

    try:
        with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            for month_data in months:
                year = month_data.get("year")
                month = month_data.get("month")
                month_total = month_data.get("total", 0)
                invoices = month_data.get("invoices", [])

                folder_name = (
                    f"{year}.{month:02d}（{len(invoices)}张发票）-¥{month_total:.2f}/"
                )

                for inv in invoices:
                    file_name = inv.get("file_name")
                    if not file_name:
                        continue

                    file_path = os.path.join(temp_dir, file_name)

                    if os.path.exists(file_path):
                        zip_file.write(file_path, folder_name + file_name)

            # 生成Excel报销单
            try:
                meta["total"] = total_amount
                meta["chinese_price"] = f"金额大写：{to_chinese_upper(total_amount)}"
                meta["total_invoice"] = len(all_invoices)
                meta["total_other_invoice"] = data.get("total_other_invoice", 0)
                excel_buf = generate_excel(
                    meta,
                    all_invoices,
                    total_amount,
                    len(all_invoices),
                    meta["total_other_invoice"],
                )
                zip_file.writestr("费用报销单.xlsx", excel_buf.read())
                print("Excel added to ZIP successfully")
            except Exception as e:
                print(f"Excel generation failed: {e}")
                import traceback

                traceback.print_exc()

        buffer.seek(0)

        filename = f"{first_year}-¥{total_amount:.2f}.zip"
        return StreamingResponse(
            buffer,
            media_type="application/zip",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
        )
    finally:
        pass


@local_invoice_routes.post("/api/download-month-pdf")
async def download_month_pdf(data: dict):
    """生成月份汇总PDF"""
    from fastapi.responses import StreamingResponse

    year = data.get("year")
    month = data.get("month")
    invoices = data.get("invoices", [])
    total = data.get("total", 0)

    # 使用reportlab生成PDF
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from io import BytesIO

    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # 标题
    c.setFont("Helvetica-Bold", 16)
    c.drawString(20 * mm, height - 20 * mm, f"{year}年{month}月发票汇总")

    # 统计信息
    c.setFont("Helvetica", 10)
    c.drawString(20 * mm, height - 35 * mm, f"发票数量: {len(invoices)}")
    c.drawString(20 * mm, height - 43 * mm, f"总金额: ¥{total:.2f}")
    c.drawString(
        20 * mm,
        height - 51 * mm,
        f"生成日期: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
    )

    # 表头
    y = height - 65 * mm
    c.setFont("Helvetica-Bold", 9)
    c.drawString(20 * mm, y, "日期")
    c.drawString(50 * mm, y, "公司")
    c.drawString(120 * mm, y, "金额")
    c.drawString(160 * mm, y, "发票号")

    # 分隔线
    y -= 3 * mm
    c.line(20 * mm, y, width - 20 * mm, y)

    # 数据行
    c.setFont("Helvetica", 8)
    y -= 5 * mm

    for i, inv in enumerate(invoices):
        if y < 30 * mm:
            c.showPage()
            y = height - 20 * mm
            c.setFont("Helvetica", 8)

        # 日期
        c.drawString(20 * mm, y, str(inv.get("date", "-")))
        # 公司（截断）
        company = inv.get("company", "-")
        if len(company) > 25:
            company = company[:25] + "..."
        c.drawString(50 * mm, y, company)
        # 金额
        c.drawString(120 * mm, y, f"¥{inv.get('amount', 0):.2f}")
        # 发票号
        invoice_no = str(inv.get("invoice_no", "-"))
        c.drawString(
            160 * mm, y, invoice_no[:15] if len(invoice_no) > 15 else invoice_no
        )

        y -= 6 * mm

    c.save()
    buffer.seek(0)

    filename = f"{year}-{month:02d}-发票汇总.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@local_invoice_routes.post("/api/download-all-pdf")
async def download_all_pdf(data: dict):
    """生成全部发票汇总PDF"""
    from fastapi.responses import StreamingResponse
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import mm
    from io import BytesIO

    months = data.get("months", [])
    total_amount = data.get("total_amount", 0)

    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # 标题
    c.setFont("Helvetica-Bold", 16)
    c.drawString(20 * mm, height - 20 * mm, "发票汇总报告")

    # 统计信息
    invoice_count = sum(len(m["invoices"]) for m in months)
    c.setFont("Helvetica", 10)
    c.drawString(20 * mm, height - 35 * mm, f"发票总数: {invoice_count}")
    c.drawString(20 * mm, height - 43 * mm, f"总金额: ¥{total_amount:.2f}")
    c.drawString(20 * mm, height - 51 * mm, f"月份数量: {len(months)}")
    c.drawString(
        20 * mm,
        height - 59 * mm,
        f"生成日期: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
    )

    y = height - 75 * mm

    for month_data in months:
        if y < 50 * mm:
            c.showPage()
            y = height - 20 * mm

        # 月份标题
        c.setFont("Helvetica-Bold", 11)
        c.drawString(
            20 * mm,
            y,
            f"{month_data['year']}年{month_data['month']}月 - 小计: ¥{month_data['total']:.2f} ({len(month_data['invoices'])}张)",
        )
        y -= 8 * mm

        # 月份内的发票
        c.setFont("Helvetica", 8)
        for inv in month_data["invoices"]:
            if y < 30 * mm:
                c.showPage()
                y = height - 20 * mm
                c.setFont("Helvetica", 8)

            company = inv.get("company", "-")
            if len(company) > 30:
                company = company[:30] + "..."

            c.drawString(
                25 * mm,
                y,
                f"{inv.get('date', '-')} | {company} | ¥{inv.get('amount', 0):.2f}",
            )
            y -= 5 * mm

        y -= 3 * mm

    c.save()
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": "attachment; filename*=UTF-8''发票汇总报告.pdf"
        },
    )
